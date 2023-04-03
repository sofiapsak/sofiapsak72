import mysql.connector
import pandas as pd
import os
import openpyxl
import xlsxwriter
import xlwt
from xlwt.Workbook import *
from pandas import ExcelWriter
 
# TEMPAT FILE
input_loc='xlsm/'
output_loc='output/'

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

files=os.listdir(input_loc)

# VARIABEL PANDAS

# PERULANGAN UNTUK PER VARIABEL
Step_1Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Step 1'):
          df1= pd.read_excel('xlsm/'+file,header=[7],sheet_name="Step 1")
          Step_1Append=Step_1Append.append(df1)

step1=Step_1Append

Step_2Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Step 2'):
          df2= pd.read_excel('xlsm/'+file,header=[8],sheet_name="Step 2")
          Step_2Append=Step_2Append.append(df2)

step2=Step_2Append

Step_3Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Step 3'):
          df3= pd.read_excel('xlsm/'+file,header=[8],sheet_name="Step 3")
          Step_3Append=Step_3Append.append(df3)

step3=Step_3Append

Step_4Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Step 4'):
          df4= pd.read_excel('xlsm/'+file,header=[7],sheet_name="Step 4")
          Step_4Append=Step_4Append.append(df4)

step4=Step_4Append

Step_5Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Step 5'):
          df5= pd.read_excel('xlsm/'+file,header=[7],sheet_name="Step 5")
          Step_5Append=Step_5Append.append(df5)

step5=Step_5Append

comp_adj_per_kontrak_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Comp Adj Per Kontrak'):
          df5= pd.read_excel('xlsm/'+file,header=[7],sheet_name="Comp Adj Per Kontrak")
          comp_adj_per_kontrak_Append=comp_adj_per_kontrak_Append.append(df5)

comp_adj_per_kontrak=comp_adj_per_kontrak_Append


comp_adj_per_gl_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Comp Adj Per GL'):
          df5= pd.read_excel('xlsm/'+file,header=[7],sheet_name="Comp Adj Per GL")
          comp_adj_per_gl_Append=comp_adj_per_gl_Append.append(df5)

comp_adj_per_gl=comp_adj_per_gl_Append

kk_4_4_q3_2022_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('KK 4.4_Q3 2022'):
          df5= pd.read_excel('xlsm/'+file,header=[9],sheet_name="KK 4.4_Q3 2022")
          kk_4_4_q3_2022_Append=kk_4_4_q3_2022_Append.append(df5)

kk_4_4_q3_2022=kk_4_4_q3_2022_Append

comp_summary_of_unsatisfied_po_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Comp Summary of Unsatisfied PO'):
          df5= pd.read_excel('xlsm/'+file,header=[8],sheet_name="Comp Summary of Unsatisfied PO")
          comp_summary_of_unsatisfied_po_Append=comp_summary_of_unsatisfied_po_Append.append(df5)

comp_summary_of_unsatisfied_po=comp_summary_of_unsatisfied_po_Append

proyeksi_revenue_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Proyeksi Revenue'):
          df5= pd.read_excel('xlsm/'+file,header=[9],sheet_name="Proyeksi Revenue")
          proyeksi_revenue_Append=proyeksi_revenue_Append.append(df5)

proyeksi_revenue=proyeksi_revenue_Append

alokasi_cacl_per_bp_number_Append=pd.DataFrame()
for file in files:
    sheets = pd.ExcelFile('xlsm/'+file)
    myWorkSheets=sheets.sheet_names
    for sheet in myWorkSheets:
      if sheet.startswith('Alokasi CACL per BP Number'):
          df5= pd.read_excel('xlsm/'+file,header=[8],sheet_name="Alokasi CACL per BP Number")
          alokasi_cacl_per_bp_number_Append=alokasi_cacl_per_bp_number_Append.append(df5)

alokasi_cacl_per_bp_number=alokasi_cacl_per_bp_number_Append

# MENJADIKAN EXCEL XLSX
with pd.ExcelWriter(output_loc+"df55.xlsx") as writer:

# MEMASUKKAN DATA PER SHEET
    step1.to_excel(writer, sheet_name="Step 1", index=False)
    step2.to_excel(writer, sheet_name="Step 2", index=False)
    step3.to_excel(writer, sheet_name="Step 3", index=False)
    step4.to_excel(writer, sheet_name="Step 4", index=False)
    step5.to_excel(writer, sheet_name="Step 5", index=False)# KOLOM STEP 5 ADA 103
    comp_adj_per_kontrak.to_excel(writer, sheet_name="Comp Adj Per Kontrak", index=False)
    comp_adj_per_gl.to_excel(writer, sheet_name="Comp Adj Per GL", index=False)
    kk_4_4_q3_2022.to_excel(writer, sheet_name="KK 4.4_Q3 2022", index=False)
    comp_summary_of_unsatisfied_po.to_excel(writer, sheet_name="Comp Summary of Unsatisfied PO", index=False)
    proyeksi_revenue.to_excel(writer, sheet_name="Proyeksi Revenue", index=False)
    alokasi_cacl_per_bp_number.to_excel(writer, sheet_name="Alokasi CACL per BP Number", index=False)

# ---------------------------------------------STEP 1--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# STEP 1
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Step 1"]

# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO step_1 (Column2, Column3, Column4, Column5, Column6, Column7, Column8, Column9, Column10, Column11, Column12, Column13, Column14, Column15, Column16, Column17, Column18, Column19, Column20, Column21, Column22, Column23, Column24, Column25, Column26, Column27, Column28, Column29, Column30, Column31, Column32, Column33, Column34, Column35, Column36, Column37, Column38, Column39, Column40, Column41, Column42, Column43, Column44, Column45, Column46, Column47, Column48, Column49) VALUES "
  sql += f"( '{sheet.cell(row, 	1	).value}', '{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}', '{sheet.cell(row, 	24	).value}', '{sheet.cell(row, 	25	).value}', '{sheet.cell(row, 	26	).value}', '{sheet.cell(row, 	27	).value}', '{sheet.cell(row, 	28	).value}', '{sheet.cell(row, 	29	).value}', '{sheet.cell(row, 	30	).value}', '{sheet.cell(row, 	31	).value}', '{sheet.cell(row, 	32	).value}', '{sheet.cell(row, 	33	).value}', '{sheet.cell(row, 	34	).value}', '{sheet.cell(row, 	35	).value}', '{sheet.cell(row, 	36	).value}', '{sheet.cell(row, 	37	).value}', '{sheet.cell(row, 	38	).value}', '{sheet.cell(row, 	39	).value}', '{sheet.cell(row, 	40	).value}', '{sheet.cell(row, 	41	).value}', '{sheet.cell(row, 	42	).value}', '{sheet.cell(row, 	43	).value}', '{sheet.cell(row, 	44	).value}', '{sheet.cell(row, 	45	).value}', '{sheet.cell(row, 	46	).value}', '{sheet.cell(row, 	50	).value}', '{sheet.cell(row, 	51	).value}')"
  cur.execute(sql)
print("DATA STEP 1 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------STEP 2--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# STEP 2
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Step 2"]
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO step_2 (Column2_2, Column2_3, Column2_4, Column2_5, Column2_6, Column2_7, Column2_8, Column2_9, Column2_10, Column2_11, Column2_12, Column2_13, Column2_14, Column2_15, Column2_16, Column2_17, Column2_18, Column2_19, Column2_20, Column2_21, Column2_22, Column2_23, Column2_24, Column2_25, Column2_26, Column2_27, Column2_28, Column2_29, Column2_30, Column2_31) VALUES "
  sql += f"('{sheet.cell(row,  1   ).value}', '{sheet.cell(row,    2   ).value}', '{sheet.cell(row,    3   ).value}', '{sheet.cell(row,    4   ).value}', '{sheet.cell(row,    5   ).value}', '{sheet.cell(row,    6   ).value}', '{sheet.cell(row,    7   ).value}', '{sheet.cell(row,    8   ).value}', '{sheet.cell(row,    9   ).value}', '{sheet.cell(row,    10  ).value}', '{sheet.cell(row,    11  ).value}', '{sheet.cell(row,    12  ).value}', '{sheet.cell(row,    13  ).value}', '{sheet.cell(row,    14  ).value}', '{sheet.cell(row,    15  ).value}', '{sheet.cell(row,    16  ).value}', '{sheet.cell(row,    17  ).value}', '{sheet.cell(row,    18  ).value}', '{sheet.cell(row,    19  ).value}', '{sheet.cell(row,    20  ).value}', '{sheet.cell(row,    21  ).value}', '{sheet.cell(row,    22  ).value}', '{sheet.cell(row,    23  ).value}', '{sheet.cell(row,    24  ).value}', '{sheet.cell(row,    25  ).value}', '{sheet.cell(row,    26  ).value}', '{sheet.cell(row,    27  ).value}', '{sheet.cell(row,    28  ).value}', '{sheet.cell(row,    29  ).value}', '{sheet.cell(row,    30  ).value}')"
  cur.execute(sql)
print(" - DATA STEP 2 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------STEP 3--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# STEP 3
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Step 3"]
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO step_3 (Column3_2, Column3_3, Column3_4, Column3_5, Column3_6, Column3_7, Column3_8, Column3_9, Column3_10, Column3_11, Column3_12, Column3_13, Column3_14, Column3_15, Column3_16, Column3_17, Column3_18, Column3_19, Column3_20, Column3_21, Column3_22, Column3_23, Column3_24, Column3_25, Column3_26, Column3_27, Column3_28, Column3_29, Column3_30, Column3_31, Column3_32, Column3_33, Column3_34, Column3_35, Column3_36, Column3_37, Column3_38) VALUES "
  sql += f"('{sheet.cell(row, 	1	).value}', '{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}', '{sheet.cell(row, 	24	).value}', '{sheet.cell(row, 	25	).value}', '{sheet.cell(row, 	26	).value}', '{sheet.cell(row, 	27	).value}', '{sheet.cell(row, 	28	).value}', '{sheet.cell(row, 	29	).value}', '{sheet.cell(row, 	30	).value}', '{sheet.cell(row, 	31	).value}', '{sheet.cell(row, 	32	).value}', '{sheet.cell(row, 	33	).value}', '{sheet.cell(row, 	34	).value}', '{sheet.cell(row, 	35	).value}', '{sheet.cell(row, 	36	).value}', '{sheet.cell(row, 	37	).value}')"
  cur.execute(sql)
print(" - DATA STEP 3 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------STEP 4--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# STEP 4
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Step 4"]
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO step_4 (Column4_2, Column4_3, Column4_4, Column4_5, Column4_6, Column4_7, Column4_8, Column4_9, Column4_10, Column4_11, Column4_12, Column4_13, Column4_14, Column4_15, Column4_16, Column4_17, Column4_18, Column4_19, Column4_20, Column4_21, Column4_22, Column4_23, Column4_24, Column4_25, Column4_26, Column4_27, Column4_28, Column4_29, Column4_30, Column4_31, Column4_32, Column4_33, Column4_34) VALUES "
  sql += f"('{sheet.cell(row, 	1	).value}', '{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}', '{sheet.cell(row, 	24	).value}', '{sheet.cell(row, 	25	).value}', '{sheet.cell(row, 	26	).value}', '{sheet.cell(row, 	27	).value}', '{sheet.cell(row, 	28	).value}', '{sheet.cell(row, 	29	).value}', '{sheet.cell(row, 	30	).value}', '{sheet.cell(row, 	31	).value}', '{sheet.cell(row, 	32	).value}', '{sheet.cell(row, 	33	).value}', '{sheet.cell(row, 	34	).value}', '{sheet.cell(row, 	35	).value}', '{sheet.cell(row, 	36	).value}', '{sheet.cell(row, 	37	).value}', '{sheet.cell(row, 	38	).value}', '{sheet.cell(row, 	39	).value}', '{sheet.cell(row, 	40	).value}', '{sheet.cell(row, 	41	).value}', '{sheet.cell(row, 	42	).value}', '{sheet.cell(row, 	43	).value}', '{sheet.cell(row, 	44	).value}', '{sheet.cell(row, 	45	).value}', '{sheet.cell(row, 	46	).value}', '{sheet.cell(row, 	47	).value}', '{sheet.cell(row, 	48	).value}')"
  cur.execute(sql)
print(" - DATA STEP 4 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------STEP 5--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# STEP 5
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Step 5"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO step_5 (Column5_2, Column5_3, Column5_4, Column5_5, Column5_6, Column5_7, Column5_8, Column5_9, Column5_10, Column5_11, Column5_12, Column5_13, Column5_14, Column5_15, Column5_16, Column5_17, Column5_18, Column5_19, Column5_20, Column5_21, Column5_22, Column5_23, Column5_24, Column5_25, Column5_26, Column5_27, Column5_28, Column5_29, Column5_30, Column5_31, Column5_32, Column5_33, Column5_34, Column5_35, Column5_36, Column5_37, Column5_38, Column5_39, Column5_40, Column5_41, Column5_42, Column5_43, Column5_44, Column5_45, Column5_46, Column5_47, Column5_48, Column5_49, Column5_50, Column5_51, Column5_52, Column5_53, Column5_54, Column5_55, Column5_56, Column5_57, Column5_58, Column5_59, Column5_60, Column5_61, Column5_62, Column5_63, Column5_64, Column5_65, Column5_66, Column5_67, Column5_68, Column5_69, Column5_70, Column5_71, Column5_72, Column5_73, Column5_74, Column5_75, Column5_76, Column5_77, Column5_78, Column5_79, Column5_80, Column5_81, Column5_82, Column5_83, Column5_84, Column5_85, Column5_86, Column5_87, Column5_88, Column5_89, Column5_90, Column5_91, Column5_92, Column5_93, Column5_94, Column5_95, Column5_96, Column5_97, Column5_98, Column5_99, Column5_100, Column5_101, Column5_102, Column5_103, Column5_104, Column5_105) VALUES "
  sql += f"('{sheet.cell(row, 	1	).value}', '{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}', '{sheet.cell(row, 	24	).value}', '{sheet.cell(row, 	25	).value}', '{sheet.cell(row, 	26	).value}', '{sheet.cell(row, 	27	).value}', '{sheet.cell(row, 	28	).value}', '{sheet.cell(row, 	29	).value}', '{sheet.cell(row, 	30	).value}', '{sheet.cell(row, 	31	).value}', '{sheet.cell(row, 	32	).value}', '{sheet.cell(row, 	33	).value}', '{sheet.cell(row, 	34	).value}', '{sheet.cell(row, 	35	).value}', '{sheet.cell(row, 	36	).value}', '{sheet.cell(row, 	37	).value}', '{sheet.cell(row, 	38	).value}', '{sheet.cell(row, 	39	).value}', '{sheet.cell(row, 	40	).value}', '{sheet.cell(row, 	41	).value}', '{sheet.cell(row, 	42	).value}', '{sheet.cell(row, 	43	).value}', '{sheet.cell(row, 	44	).value}', '{sheet.cell(row, 	45	).value}', '{sheet.cell(row, 	46	).value}', '{sheet.cell(row, 	47	).value}', '{sheet.cell(row, 	48	).value}', '{sheet.cell(row, 	49	).value}', '{sheet.cell(row, 	50	).value}', '{sheet.cell(row, 	51	).value}', '{sheet.cell(row, 	52	).value}', '{sheet.cell(row, 	53	).value}', '{sheet.cell(row, 	54	).value}', '{sheet.cell(row, 	55	).value}', '{sheet.cell(row, 	56	).value}', '{sheet.cell(row, 	57	).value}', '{sheet.cell(row, 	58	).value}', '{sheet.cell(row, 	59	).value}', '{sheet.cell(row, 	60	).value}', '{sheet.cell(row, 	61	).value}', '{sheet.cell(row, 	62	).value}', '{sheet.cell(row, 	63	).value}', '{sheet.cell(row, 	64	).value}', '{sheet.cell(row, 	65	).value}', '{sheet.cell(row, 	66	).value}', '{sheet.cell(row, 	67	).value}', '{sheet.cell(row, 	68	).value}', '{sheet.cell(row, 	69	).value}', '{sheet.cell(row, 	70	).value}', '{sheet.cell(row, 	71	).value}', '{sheet.cell(row, 	72	).value}', '{sheet.cell(row, 	73	).value}', '{sheet.cell(row, 	74	).value}', '{sheet.cell(row, 	75	).value}', '{sheet.cell(row, 	162	).value}', '{sheet.cell(row, 	163	).value}', '{sheet.cell(row, 	164	).value}', '{sheet.cell(row, 	165	).value}', '{sheet.cell(row, 	168	).value}', '{sheet.cell(row, 	175	).value}', '{sheet.cell(row, 	176	).value}', '{sheet.cell(row, 	177	).value}', '{sheet.cell(row, 	178	).value}', '{sheet.cell(row, 	179	).value}', '{sheet.cell(row, 	180	).value}', '{sheet.cell(row, 	181	).value}', '{sheet.cell(row, 	182	).value}', '{sheet.cell(row, 	183	).value}', '{sheet.cell(row, 	184	).value}', '{sheet.cell(row, 	185	).value}', '{sheet.cell(row, 	186	).value}', '{sheet.cell(row, 	187	).value}', '{sheet.cell(row, 	188	).value}', '{sheet.cell(row, 	231	).value}', '{sheet.cell(row, 	232	).value}', '{sheet.cell(row, 	233	).value}', '{sheet.cell(row, 	234	).value}', '{sheet.cell(row, 	235	).value}', '{sheet.cell(row, 	236	).value}', '{sheet.cell(row, 	237	).value}', '{sheet.cell(row, 	238	).value}', '{sheet.cell(row, 	239	).value}', '{sheet.cell(row, 	240	).value}')"
  cur.execute(sql)
print(" - DATA STEP 5 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------Comp Adj Per Kontrak--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# Comp Adj Per Kontrak
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Comp Adj Per Kontrak"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO comp_adj_per_kontrak (Column6_2, Column6_3, Column6_4, Column6_5, Column6_6, Column6_7, Column6_8, Column6_9, Column6_10, Column6_11, Column6_12, Column6_13, Column6_14, Column6_15, Column6_16, Column6_17, Column6_18, Column6_19, Column6_20, Column6_21, Column6_22, Column6_23) VALUES "
  sql += f"('{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}')"
  cur.execute(sql)
print(" - DATA STEP Comp Adj Per Kontrak BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------Comp Adj Per GL--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# Comp Adj Per GL
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Comp Adj Per GL"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO comp_adj_per_gl (Column7_2, Column7_3, Column7_4, Column7_5, Column7_6, Column7_7, Column7_8, Column7_9, Column7_10, Column7_11, Column7_12, Column7_13, Column7_14, Column7_15, Column7_16, Column7_17, Column7_18, Column7_19, Column7_20, Column7_21, Column7_22) VALUES "
  sql += f"('{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}')"
  cur.execute(sql)
print(" - DATA STEP Comp Adj Per GL BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------KK 4.4_Q3 2022--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# KK 4.4_Q3 2022
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["KK 4.4_Q3 2022"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO kk_44 (Column8_2, Column8_3, Column8_4, Column8_5, Column8_6, Column8_7, Column8_8, Column8_9, Column8_10, Column8_11, Column8_12, Column8_13, Column8_14, Column8_15, Column8_16, Column8_17, Column8_18, Column8_19) VALUES "
  sql += f"('{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}')"
  cur.execute(sql)
print(" - DATA STEP KK 4.4 BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------Comp Summary of Unsatisfied PO--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# Comp Summary of Unsatisfied PO
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Comp Summary of Unsatisfied PO"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO comp_summary_of_unsatisfied_po (Column9_2, Column9_3, Column9_4, Column9_5, Column9_6, Column9_7, Column9_8, Column9_9, Column9_10, Column9_11, Column9_12, Column9_13, Column9_14, Column9_15, Column9_16, Column9_17) VALUES "
  sql += f"('{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}')"
  cur.execute(sql)
print(" - DATA STEP Comp Summary of Unsatisfied PO BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------Proyeksi Revenue--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# Proyeksi Revenue
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Proyeksi Revenue"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO proyeksi_revenue (Column10_2, Column10_3, Column10_4, Column10_5, Column10_6, Column10_7, Column10_8, Column10_9, Column10_10, Column10_11) VALUES "
  sql += f"('{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}')"
  cur.execute(sql)
print(" - DATA STEP Proyeksi Revenue BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# ---------------------------------------------Alokasi CACL per BP Number--------------------------------------------------
# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()

# Alokasi CACL per BP Number
# (B) OPEN EXCEL FILE OUTPUT
book = openpyxl.load_workbook("output/df55.xlsx")
sheet = book["Alokasi CACL per BP Number"]
print(sheet)
# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
  sql = "INSERT INTO alokasi_cacl_per_bp_number (Column11_2, Column11_3, Column11_4, Column11_5, Column11_6, Column11_7, Column11_8, Column11_9, Column11_10, Column11_11, Column11_12, Column11_13, Column11_14, Column11_15, Column11_16, Column11_17, Column11_18, Column11_19, Column11_20, Column11_21, Column11_22, Column11_23, Column11_24, Column11_25, Column11_26) VALUES "
  sql += f"('{sheet.cell(row, 	1	).value}', '{sheet.cell(row, 	2	).value}', '{sheet.cell(row, 	3	).value}', '{sheet.cell(row, 	4	).value}', '{sheet.cell(row, 	5	).value}', '{sheet.cell(row, 	6	).value}', '{sheet.cell(row, 	7	).value}', '{sheet.cell(row, 	8	).value}', '{sheet.cell(row, 	9	).value}', '{sheet.cell(row, 	10	).value}', '{sheet.cell(row, 	11	).value}', '{sheet.cell(row, 	12	).value}', '{sheet.cell(row, 	13	).value}', '{sheet.cell(row, 	14	).value}', '{sheet.cell(row, 	15	).value}', '{sheet.cell(row, 	16	).value}', '{sheet.cell(row, 	17	).value}', '{sheet.cell(row, 	18	).value}', '{sheet.cell(row, 	19	).value}', '{sheet.cell(row, 	20	).value}', '{sheet.cell(row, 	21	).value}', '{sheet.cell(row, 	22	).value}', '{sheet.cell(row, 	23	).value}', '{sheet.cell(row, 	24	).value}', '{sheet.cell(row, 	25	).value}')"
  cur.execute(sql)
print(" - DATA STEP Alokasi CACL per BP Number BERSHASIL DITAMBAHKAN")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM step_1 WHERE Column49='Pending'"
cur.execute(sql)
print(" - DATA STEP 1 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM step_2 WHERE Column2_31='Pending'"
cur.execute(sql)
print(" - DATA STEP 2 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM step_3 WHERE Column3_38='Pending'"
cur.execute(sql)
print(" - DATA STEP 3 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM step_4 WHERE Column4_34='Pending'"
cur.execute(sql)
print(" - DATA STEP 4 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM step_5 WHERE Column5_105='Pending'"
cur.execute(sql)
print(" - DATA STEP 5 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM comp_adj_per_kontrak WHERE Column6_23='Pending'"
cur.execute(sql)
print(" - DATA comp_adj_per_kontrak BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM comp_adj_per_gl WHERE Column7_22='Pending'"
cur.execute(sql)
print(" - DATA comp_adj_per_gl BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM kk_44 WHERE Column8_19='Pending'"
cur.execute(sql)
print(" - DATA kk_44 BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM comp_summary_of_unsatisfied_po WHERE Column9_17='Pending'"
cur.execute(sql)
print(" - DATA comp_summary_of_unsatisfied_po BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM proyeksi_revenue WHERE Column10_11='Pending'"
cur.execute(sql)
print(" - DATA proyeksi_revenue BERSHASIL DIHAPUS")
conn.commit()
conn.close()

# (A) CONNECT TO DATABASE
conn=mysql.connector.connect(host="localhost", user="root", passwd="", database="bdo")
cur=conn.cursor()
# (B) DELETE DATA PENDING
sql = "DELETE FROM alokasi_cacl_per_bp_number WHERE Column11_26='Pending'"
cur.execute(sql)
print(" - DATA alokasi_cacl_per_bp_number BERSHASIL DIHAPUS")
conn.commit()
conn.close()